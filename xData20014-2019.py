# Extract the failure info from the disparate daily report stored in individual excel files.

from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import os
import re

##--- Script Settings
filePath = 'data'   # Path of the directory with the daily reports

# Get the file list
fileList = os.listdir(filePath)
numFiles = len(fileList)

# Generate the date range
dateList = pd.date_range(datetime(2014,8,14), datetime(2020,1,1), freq='D')

# Generate the empty dataframe for raw failure info
rawData2014To2019 = pd.Series(index=dateList, name='failInfo')


# -- Extract the data into the dataframe
ii = 0
for fileName in fileList:
    print('Files left:' + str(len(fileList) - ii) + '/' + str(numFiles))
    ii += 1
    try:
        workbook = load_workbook(os.path.join(filePath, fileName), data_only=True)
    except Exception:
        msg = fileName + ' - Failed to read the file. Skipped.'
        print(msg)
        continue

    try:
        sheet = workbook['daily report']
    except KeyError:
        try:
            sheet = workbook['Daily Report']
        except KeyError:
            msg = fileName + ' - Daily report not found. Skipped.'
            print(msg)
            continue

    # Extract the date from the fileName (more reliable than the date in the report)
    dateRx = re.compile(r'(\d{1,2}-\d{1,2}-\d{4})')
    dateStr = dateRx.findall(fileName)[0]
    reportDate = datetime.strptime(dateStr, '%m-%d-%Y')

    # Gather the data from the ROI into a string
    dataStr= ''
    for row in sheet.iter_rows(min_row=20, max_col=8, values_only=True):
        dataStr += str(row)
        # Preprocessing
        # Step 1. Remove all the infomation pertaining to (following) N. Coleman
        dataStr = re.sub(re.compile(r'N\.\s?Coleman.*'), '', dataStr)
        # Step 2. Remove "None", datetimes and parentheses
        dataStr = re.sub(re.compile(r'None|\)|\(|,|datetime\.datetime\(.*?\)|#|Well|Down List|Status|Date '
                                    r'Down|Est. Oil|Coleman|\d{1,2}/\d{1,2}/\d{2,4}|loaded|tested|Problem'), '',
                         dataStr)
        # Step 3. Remove the stand-alone digits.
        dataStr = re.sub(re.compile(r'(?<= )\d{1,2}(?= )'), '', dataStr)
        # Step 4. Remove the total production loss digits from the end.
        dataStr = re.sub(re.compile(r'\d{1,3}\s+$'), '', dataStr)


    rawData2014To2019.loc[reportDate] = dataStr

# Write actMatrix to csv file
rawData2014To2019.to_csv('rawData2_2014-2019.csv')